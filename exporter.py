import sqlite3

from openpyxl import Workbook
from access_parser import AccessParser

EXCEL_MAX_ROWS = 1_048_576


def safe_value(value):
    if isinstance(value, bytes):
        return "0x" + value.hex().upper()
    return value


def safe_sheet_name(name):
    for ch in '[]:*?/\\':
        name = name.replace(ch, "_")
    return name[:31] or "Tabell"


def unique_sheet_name(name, used):
    base = safe_sheet_name(name)
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = base[:31-len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def sqlite_connect_ro(path):
    return sqlite3.connect(f"file:{path}?mode=ro", uri=True)


def detect_database(path, ext):
    ext = ext.lower()

    if ext in {".db3", ".sqlite", ".sqlite3"}:
        try:
            with sqlite_connect_ro(path) as conn:
                tables = [r[0] for r in conn.execute(
                    "SELECT name FROM sqlite_master "
                    "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                ).fetchall()]
            return {"kind": "sqlite", "label": "SQLite / DB3", "tables": len(tables)}
        except Exception:
            pass

    if ext in {".mdb", ".accdb"}:
        try:
            db = AccessParser(path)
            tables = [name for name in db.catalog if not name.startswith("MSys")]
            return {"kind": "access", "label": "Access / PEX", "tables": len(tables)}
        except Exception:
            pass

    try:
        with sqlite_connect_ro(path) as conn:
            tables = [r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master "
                "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()]
        return {"kind": "sqlite", "label": "SQLite / DB3", "tables": len(tables)}
    except Exception:
        pass

    try:
        db = AccessParser(path)
        tables = [name for name in db.catalog if not name.startswith("MSys")]
        return {"kind": "access", "label": "Access / PEX", "tables": len(tables)}
    except Exception as exc:
        raise ValueError("Databasformatet kunde inte identifieras eller läsas.") from exc


def export_sqlite(input_path, output_path, progress):
    with sqlite_connect_ro(input_path) as conn:
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()]

        wb = Workbook(write_only=True)
        used = set()
        total = len(tables)
        last_table = ""
        last_rows = 0

        for index, table in enumerate(tables, start=1):
            last_table = table
            rows_written = 0
            safe_table = table.replace('"', '""')
            cur = conn.execute(f'SELECT * FROM "{safe_table}"')
            columns = [d[0] for d in cur.description]

            sheet_number = 1
            ws = wb.create_sheet(unique_sheet_name(table, used))
            ws.append(columns)
            rows_in_sheet = 1

            while True:
                batch = cur.fetchmany(5000)
                if not batch:
                    break
                for row in batch:
                    if rows_in_sheet >= EXCEL_MAX_ROWS:
                        sheet_number += 1
                        ws = wb.create_sheet(unique_sheet_name(f"{table}_{sheet_number}", used))
                        ws.append(columns)
                        rows_in_sheet = 1
                    ws.append([safe_value(v) for v in row])
                    rows_in_sheet += 1
                    rows_written += 1
                progress(index, total, table, rows_written)

            last_rows = rows_written
            progress(index, total, table, rows_written)

        wb.save(output_path)
        return {"tables": total, "last_table": last_table, "last_rows": last_rows}


def export_access(input_path, output_path, progress):
    db = AccessParser(input_path)
    tables = [name for name in db.catalog if not name.startswith("MSys")]

    wb = Workbook(write_only=True)
    used = set()
    total = len(tables)
    last_table = ""
    last_rows = 0

    for index, table_name in enumerate(tables, start=1):
        last_table = table_name
        table = db.parse_table(table_name)
        columns = list(table.keys())
        row_count = len(table[columns[0]]) if columns else 0

        sheet_number = 1
        ws = wb.create_sheet(unique_sheet_name(table_name, used))
        ws.append(columns)
        rows_in_sheet = 1

        for row_index in range(row_count):
            if rows_in_sheet >= EXCEL_MAX_ROWS:
                sheet_number += 1
                ws = wb.create_sheet(unique_sheet_name(f"{table_name}_{sheet_number}", used))
                ws.append(columns)
                rows_in_sheet = 1

            ws.append([safe_value(table[col][row_index]) for col in columns])
            rows_in_sheet += 1

            if row_index and row_index % 5000 == 0:
                progress(index, total, table_name, row_index + 1)

        last_rows = row_count
        progress(index, total, table_name, row_count)

    wb.save(output_path)
    return {"tables": total, "last_table": last_table, "last_rows": last_rows}


def export_database(input_path, output_path, kind, progress):
    if kind == "sqlite":
        return export_sqlite(input_path, output_path, progress)
    if kind == "access":
        return export_access(input_path, output_path, progress)
    raise ValueError(f"Okänd databastyp: {kind}")
